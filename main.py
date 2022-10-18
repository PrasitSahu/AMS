import mysql.connector as mysql
from env import password
from datetime import date
import os
from os.path import exists
from openpyxl import Workbook


def cls():
    """ clears the console's screen """

    os.system('cls' if os.name == 'nt' else 'clear')


months = ['january', 'february', 'march', 'april', 'may', 'june',
          'july', 'august', 'september', 'october', 'november', 'december']


def record_attendance():
    """ Adds new records to the attendance table/register """

    # get today's date
    today = date.today()
    this_year = today.year
    this_month = today.month
    this_day = today.day

    # connection to students main database
    connection = connect_db('classxii')
    cursor = connection.cursor(buffered=True)
    table = f"{months[this_month - 1]}_{this_year}"

    # gets all the tables from database
    cmd = f"SHOW TABLES;"
    cursor.execute(cmd)
    tables = []
    for i in cursor:
        tables.append(i[0])

    if table not in tables:
        cmd = f"CREATE TABLE {table}(" \
              f"id int NOT NULL AUTO_INCREMENT PRIMARY KEY," \
              f"name varchar(60) NOT NULL" \
              f");"
        # add tables
        cursor.execute(cmd)
        connection.commit()

        # add students data to the new table
        cmd = f"INSERT INTO {table}(id, name) SELECT id, name FROM attendance;"
        cursor.execute(cmd)
        connection.commit()

    cmd1 = f'SELECT * FROM {table};'
    cmd2 = 'SELECT * FROM attendance;'
    cursor.execute(cmd1)
    # stores all the date columns from the database
    prev_dates = cursor.column_names[2:]
    col = f"{this_day}d_{this_month}m_{this_year}"
    remove_and_add = False
    old_col = col
    if col in prev_dates:
        print('Warn: Attendance record for today is already available!\n')
        res = input('Do you want to erase the current record and write a new one? ')
        if res.lower() not in ['n', 'no']:
            col = None
            remove_and_add = True

    if col not in prev_dates:
        cls()
        record = []
        student_names = []
        student_ids = []

        print(f'''Record for: {this_day}/ {this_month}/ {this_year}\n''')
        print("Note: Type 'q' or 'quit' to exit at anytime while entering the records (No entered data wil be saved)\n")
        cursor.execute(cmd2)
        break_outer = False
        for student in cursor:
            while True:
                prompt = input(f"{student[0]}. {student[1]} - ")
                student_names.append(student[1])
                student_ids.append(student[0])
                if prompt.lower() in ['y', 'yes']:
                    record.append('present')
                    break
                elif prompt.lower() in ['n', 'no']:
                    record.append('absent')
                    break
                elif prompt.lower() in ['q', 'quit']:
                    break_outer = True
                    break
                else:
                    print('\nInvalid Input, Try again...')
                    continue
            if break_outer:
                break

        if not break_outer:
            if remove_and_add:
                # sql command to delete a date column from the table
                cmd = f"ALTER TABLE {table} DROP COLUMN {old_col};"
                cursor.execute(cmd)
                connection.commit()

            # sql command to add new date column to the table
            cmd = f"ALTER TABLE {table} ADD {old_col} varchar(7)"

            # add new date column
            cursor.execute(cmd)
            connection.commit()

            try:
                # sql command to get data from a particular table
                cmd = f"SELECT * FROM {table}"
                cursor.execute(cmd)
                count = 0
                for st in cursor:
                    count += 1

                if count != len(record):
                    for j in range(count, len(record)):
                        cmd = f"INSERT INTO {table}(id, name) VALUES({student_ids[j]}, '{student_names[j]}')"
                        cursor.execute(cmd)
                        connection.commit()
                for i in range(len(record)):
                    cmd = f"UPDATE {table} SET {old_col} = '{record[i]}' WHERE id = {i + 1};"
                    cursor.execute(cmd)
                    connection.commit()

                print('\nRecord Updated Successfully. :) \n')

                while True:
                    to_update_excel = input('want to update in excel sheet? ')
                    if to_update_excel.lower() in ['yes', 'y']:
                        update_records()
                        print("Updated Successfully :)")
                        break
                    elif to_update_excel.lower() in ['no', 'n']:
                        break
                    else:
                        print('\nInvalid input, Try again...')
                        continue

            except mysql.Error as error:
                # update failed message as an error
                print(f"\nRecord Update Failed !: {error}")

                # reverting changes because of exception
                connection.rollback()


def update_records():
    wb = Workbook()
    ws = wb.active

    # connection to students main database
    connection = connect_db('classxii')
    cursor = connection.cursor()

    cmd = "SHOW TABLES;"
    cursor.execute(cmd)
    cls()
    print("select a table - ")
    num = 1
    tables = []
    for table in cursor:
        if table[0].lower() == 'attendance':
            continue
        print(f"{num}. {table[0]}")
        tables.append(table[0])
        num += 1
    table_num = None
    while True:
        try:
            table_num = int(input('\n> '))
            if table_num in range(1, len(tables) + 1):
                break
            else:
                print('Invalid input, try again...')
                continue
        except ValueError:
            print('Invalid input, try again...')
            continue
    cursor.execute(f'SELECT * FROM {tables[table_num - 1]};')

    # stores all the columns from the database
    columns = list(cursor.column_names[:])
    columns.append('Percentage')
    row = 1
    col = 0
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
            'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE']
    for column in columns:
        ws[f"{cols[col]}{row}"] = column
        col += 1

    row = 2
    col = 0
    for student in cursor:
        p_count = 0
        for i in student[2:]:
            if i == 'present':
                p_count += 1
        percent = round((p_count / len(student[2:]) * 100), 2)
        student = list(student)
        student.append(f'{percent}%')
        for info in student:
            if info is None:
                info = '  -  '
            ws[f"{cols[col].strip()}{row}"] = info
            col += 1
        row += 1
        col = 0

    file = f'{tables[table_num - 1]}.xlsx'
    if exists(file):
        os.remove(file)
    wb.save(file)
    print('updated successfully. :)\n')


def add_student():
    """ adds a new student to the table """

    connection = connect_db('classxii')
    cursor = connection.cursor()
    cmd = 'SELECT * FROM attendance;'
    cursor.execute(cmd)
    num = 0
    for i in cursor:
        num += 1
    num += 1
    c_num = num
    nms = []
    print("Note: Type 'q' or 'quit' to quit after entering all the names\n")
    print('Enter the names: ')

    while True:
        name_inp = input(f"{num}: ")
        name_inp.strip()
        if name_inp.lower() in ['q', 'quit']:
            break
        check = input('you sure: ')
        print()
        if check.lower() in ['n', 'no']:
            print(f'## {name_inp} not added ##\n')
            continue
        if check.lower() in ['q', 'quit']:
            break
        nms.append(name_inp)
        num += 1
    for name in nms:
        cmd = f"INSERT INTO attendance(id, name) VALUES({c_num},'{name}')"
        cursor.execute(cmd)
        connection.commit()
        c_num += 1

    if len(nms):
        print('\n Data Saved Successfully.\n')
    else:
        print('\n No data was saved.\n')


def show_menu():
    print('''
                                            ATTENDANCE MANAGEMENT SYSTEM - AMS

            SELECT AN OPTION - 
        ────────────────────────
        1. RECORD ATTENDANCE
        2. UPDATE IN EXCEL SHEET
        3. ADD STUDENT
        4. HOW TO ??
        5. Quit

            ''')


def connect_db(db_name):
    """ makes a connection to mysql db """

    connection = mysql.connect(host='localhost',
                               user='root',
                               password=password or '',
                               database=db_name)
    return connection


def main():
    """ Main (Root) Function """

    cls()
    show_menu()

    while True:
        try:
            opt = int(input('  > '))
            if opt not in [1, 2, 3, 4, 5]:
                print('Invalid input! try again...')
                continue
            elif opt == 5:
                print('\nBye! Bye!\n')
                break
            else:
                cls()
                if opt == 1:
                    record_attendance()
                    input('go back ↩ ')
                    cls()
                    show_menu()
                    continue
                elif opt == 2:
                    update_records()
                    input('go back ↩ ')
                    cls()
                    show_menu()
                    continue
                elif opt == 3:
                    add_student()
                    input('go back ↩ ')
                    cls()
                    show_menu()
                    continue
                else:
                    print('''
                                                                HOW TO USE
                                                               ────────────
                        It is a very useful tool for maintaining an attendance record. The interface is very easy to understand.
                        To mark as 'Present' just type 'Yes' or 'Y' and to mark as 'Absent' just type 'No' or 'N'. 
                                  You can even save the records as an excel sheet by choosing the 'UPDATE IN EXCEL SHEET' option
                        from the main menu. 
                    ''')
                    input('go back ↩ ')
                    cls()
                    show_menu()
                    continue
        except ValueError:
            print(ValueError)
            continue


if __name__ == '__main__':
    main()
